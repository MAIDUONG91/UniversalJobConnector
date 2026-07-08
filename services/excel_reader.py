from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from models.filter_config import FilterConfig


class ExcelReader:

    def __init__(self):

        self.file = Path("config/filter_config.xlsx")

    def load(self) -> FilterConfig:

        workbook = load_workbook(self.file)

        sheet = workbook["Filter"]

        data = {}

        for key, value in sheet.iter_rows(
            min_row=1,
            max_col=2,
            values_only=True
        ):

            data[key] = value

        workbook.close()

        return FilterConfig(
            keyword=data["keyword"],
            location=data["location"],
            min_salary=int(data["min_salary"]),
            max_salary=int(data["max_salary"]),
        )